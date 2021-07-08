# 토지 임야 정보 api 가져오기(XML)
from bs4 import BeautifulSoup
from dataclasses import dataclass
import openpyxl
import requests
import time
import heapq

# 다른 정보를 검색하고 싶을때 아래 코드를 바꿔 주시면 됩니다.
# 여기에 신청한 본인의 key를 넣어야 합니다
authKey = '9d0d9e0f456f80711f7d8e'
# 법정동 코드
pnu = '4146311000'
# 사이트 주소
url = 'http://openapi.nsdi.go.kr/nsdi/eios/LadfrlService/ladfrlList.xml'
#-----------------------------------------------------------------------


numOfRows = '10'
Result = []
@dataclass
class Data:
    # 지번
    mnnmSlno : str
    # 지목
    lndcgrCodeNm : str
    # 면적
    lndpclAr : str
    # 소유구분
    posesnSeCodeNm : str
    # 데이터 기준 일자
    lastUpdtDt : str


def inputData():
    authKey = input('키를 넣어주세요 : ')
    pnu = input('법정동 코드를 넣어주세요 : ')

# 데이터를 가져오는 함수
def getCityData():
    # 페이지
    pageNo = 0
    while(1):
        pageNo+=1
        queryParams = '?'+'authkey='+authKey+'&pnu='+pnu+'&numOfRows='+numOfRows+'&pageNo='+str(pageNo)
        req = requests.get(url+queryParams)
        print(url+queryParams)
        html = req.text
        soup = BeautifulSoup(html,'html.parser')
        my_contnet = soup.select('fields > ladfrlVOList')
        if(len(my_contnet)==0):
            break
        for content in my_contnet:
            try:
                data = Data(
                    mnnmSlno = content.select_one('mnnmSlno').get_text(),
                    lndcgrCodeNm = content.select_one('lndcgrCodeNm').get_text(),
                    lndpclAr = content.select_one('lndpclAr').get_text(),
                    posesnSeCodeNm = content.select_one('posesnSeCodeNm').get_text(),
                    lastUpdtDt = content.select_one('lastUpdtDt').get_text()
                )
            # 소유구분이 없는 데이터가 있어서 예외처리 함
            except AttributeError as e:
                data = Data(
                    mnnmSlno = content.select_one('mnnmSlno').get_text(),
                    lndcgrCodeNm = content.select_one('lndcgrCodeNm').get_text(),
                    lndpclAr = content.select_one('lndpclAr').get_text(),
                    posesnSeCodeNm = '',
                    lastUpdtDt = content.select_one('lastUpdtDt').get_text()
                )

            Result.append(data)

# 엑셀에 기입하는 함수
def excelInput():
    wb = openpyxl.Workbook()
    sheet1 = wb.active
    for i in range(1,len(Result)):
        sheet1.cell(row = i,column = 1,value = Result[i].mnnmSlno)
        sheet1.cell(row = i,column = 2,value = Result[i].lndcgrCodeNm)
        sheet1.cell(row = i,column = 3,value = Result[i].lndpclAr)
        sheet1.cell(row = i,column = 4,value = Result[i].posesnSeCodeNm)
        sheet1.cell(row = i,column = 5,value = Result[i].lastUpdtDt)
    wb.save('토지임야정보.xlsx')

def main():
    #inputData()
    getCityData()
    excelInput()
main()